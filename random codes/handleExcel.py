from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from typing import Optional


def getColumn(file: str, column: int, rows: Optional[int] = None) -> Cell:
    # get the contents of a column
    wb: Workbook = load_workbook(filename=file)
    # grab the active worksheet
    sheet = wb.active
    for cell in sheet.iter_cols(min_col=column, max_col=column, max_row=rows):
        return cell


for cell in getColumn("2021 combines.xlsx", 2):
    if cell.value is not None:
        print(cell.value)
